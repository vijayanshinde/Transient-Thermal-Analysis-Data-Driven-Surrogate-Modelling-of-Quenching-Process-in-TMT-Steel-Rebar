"""
validate_against_excel.py
=========================
Cell-for-cell verification of `quench_solver.py` against the source
workbook `Transient_Explicit_FEA_Fundamental_25mm_bar_example.xlsx`.

Both simulation blocks of Sheet1 are compared across all 21 nodes and
all 228 time levels. The script reports the maximum and RMS deviation
per block and writes a validation figure.
"""

from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from quench_solver import config_part1, config_part2, solve

# WORKBOOK = Path(
#     "Transient_Explicit_FEA_Fundamental_25mm_bar_example.xlsx"
# )
# OUTDIR = Path(".")

HERE = Path(__file__).parent
WORKBOOK = HERE / "Transient_Explicit_FEA_Fundamental_25mm_bar_example.xlsx"
OUTDIR = HERE
OUTDIR.mkdir(parents=True, exist_ok=True)

# Sheet1 layout: node i temperature sits in column 4 + 5*i (1-indexed).
FIRST_T_COLUMN = 4
COLUMNS_PER_NODE = 5
N_NODES = 21
N_ROWS = 228

BLOCKS = {
    "Part 1 (no radiation)": 19,    # first data row of block 1
    "Part 2 (with radiation)": 267,  # first data row of block 2
}


def read_block(sheet, first_row: int) -> np.ndarray:
    """Extract a [N_ROWS, N_NODES] temperature array from the worksheet."""
    out = np.full((N_ROWS, N_NODES), np.nan)
    for r in range(N_ROWS):
        for i in range(N_NODES):
            value = sheet.cell(
                row=first_row + r, column=FIRST_T_COLUMN + COLUMNS_PER_NODE * i
            ).value
            if isinstance(value, (int, float)):
                out[r, i] = float(value)
    return out


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook["Sheet1"]

    runs = {
        "Part 1 (no radiation)": solve(config_part1()),
        "Part 2 (with radiation)": solve(config_part2()),
    }

    print("=" * 68)
    print("VALIDATION: Python solver vs Excel workbook")
    print("=" * 68)

    excel_data = {}
    for label, first_row in BLOCKS.items():
        excel = read_block(sheet, first_row)
        excel_data[label] = excel
        python = runs[label].T

        difference = python - excel
        max_abs = np.nanmax(np.abs(difference))
        rms = float(np.sqrt(np.nanmean(difference ** 2)))
        worst = np.unravel_index(np.nanargmax(np.abs(difference)), difference.shape)

        print(f"\n{label}")
        print(f"  cells compared        : {np.sum(~np.isnan(excel)):,}")
        print(f"  max abs deviation     : {max_abs:.3e} degC")
        print(f"  RMS deviation         : {rms:.3e} degC")
        print(f"  worst cell            : step {worst[0]}, node {worst[1]}")
        print(f"  surface  Excel {excel[-1, -1]:8.3f}  Python {python[-1, -1]:8.3f}")
        print(f"  core     Excel {excel[-1, 0]:8.3f}  Python {python[-1, 0]:8.3f}")
        verdict = "PASS" if max_abs < 1e-6 else ("PASS (loose)" if max_abs < 0.01 else "FAIL")
        print(f"  verdict               : {verdict}")

    _figure(runs, excel_data)
    print(f"\nFigure written to {OUTDIR / 'validation_python_vs_excel.png'}")


def _figure(runs, excel_data) -> None:
    fig, axes = plt.subplots(1, 3, figsize=(16, 5))

    # -- (a) surface and core histories, both blocks --------------------
    ax = axes[0]
    colours = {"Part 1 (no radiation)": "#1f77b4", "Part 2 (with radiation)": "#d62728"}
    for label, result in runs.items():
        colour = colours[label]
        ax.plot(result.time, result.T_surface, color=colour, lw=2,
                label=f"{label} - surface (Python)")
        ax.plot(result.time, result.T_core, color=colour, lw=2, ls=":",
                label=f"{label} - core (Python)")
        excel = excel_data[label]
        ax.plot(result.time[::12], excel[::12, -1], "o", color=colour,
                ms=4, mfc="none", label=f"{label} - surface (Excel)")
    ax.set_xlabel("Time [s]")
    ax.set_ylabel("Temperature [degC]")
    ax.set_title("(a) Validation: Python lines vs Excel markers")
    ax.legend(fontsize=7, loc="center right")
    ax.grid(alpha=0.3)

    # -- (b) radial profiles at several instants ------------------------
    ax = axes[1]
    result = runs["Part 2 (with radiation)"]
    radius_mm = result.config.r * 1e3
    for step in (0, 25, 60, 120, 227):
        ax.plot(radius_mm, result.T[step], marker="o", ms=3,
                label=f"t = {result.time[step]:.3f} s")
    ax.set_xlabel("Radial position [mm]")
    ax.set_ylabel("Temperature [degC]")
    ax.set_title("(b) Radial profiles, Part 2")
    ax.legend(fontsize=8)
    ax.grid(alpha=0.3)

    # -- (c) deviation map ----------------------------------------------
    ax = axes[2]
    difference = runs["Part 2 (with radiation)"].T - excel_data["Part 2 (with radiation)"]
    image = ax.imshow(np.abs(difference).T, aspect="auto", origin="lower",
                      cmap="viridis",
                      extent=[0, result.time[-1], 0, N_NODES - 1])
    ax.set_xlabel("Time [s]")
    ax.set_ylabel("Node index")
    ax.set_title("(c) |Python - Excel| [degC], Part 2")
    fig.colorbar(image, ax=ax)

    fig.tight_layout()
    fig.savefig(OUTDIR / "validation_python_vs_excel.png", dpi=150)
    plt.close(fig)


if __name__ == "__main__":
    main()
